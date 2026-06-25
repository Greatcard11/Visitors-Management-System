"""
Report generation: Excel, CSV, and PDF summary reports.
"""

import io
import pandas as pd
from datetime import datetime
import pytz

LAGOS_TZ = pytz.timezone("Africa/Lagos")


def visitors_to_dataframe(visitors: list[dict]) -> pd.DataFrame:
    cols = [
        "visitor_number", "full_name", "phone", "email", "gender",
        "company", "department", "person_to_visit", "purpose",
        "id_type", "id_number", "status",
        "entry_date", "entry_time", "exit_date", "exit_time", "visit_duration",
        "expected_duration", "vehicle_reg", "items_carried",
        "emergency_name", "emergency_phone", "barcode_id"
    ]
    df = pd.DataFrame(visitors)
    for col in cols:
        if col not in df.columns:
            df[col] = ""
    return df[cols].fillna("")


def generate_excel_report(visitors: list[dict], title="Visitor Report") -> bytes:
    df = visitors_to_dataframe(visitors)
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Visitors")
        ws = writer.sheets["Visitors"]
        # Style header row
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    return buf.getvalue()


def generate_csv_report(visitors: list[dict]) -> bytes:
    df = visitors_to_dataframe(visitors)
    return df.to_csv(index=False).encode()


def generate_pdf_report(visitors: list[dict], title="Visitor Report") -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    navy = colors.HexColor("#1a2744")
    green = colors.HexColor("#22c55e")

    story = []

    # Title
    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=16,
                                 textColor=navy, alignment=TA_CENTER, spaceAfter=4)
    now_str = datetime.now(LAGOS_TZ).strftime("%d %b %Y %H:%M WAT")
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(f"Generated: {now_str}  •  Total Records: {len(visitors)}",
                            ParagraphStyle("s", parent=styles["Normal"], fontSize=9,
                                           textColor=colors.grey, alignment=TA_CENTER)))
    story.append(Spacer(1, 6*mm))

    if not visitors:
        story.append(Paragraph("No records found.", styles["Normal"]))
        doc.build(story)
        return buf.getvalue()

    # Table
    headers = ["#", "Visitor No", "Full Name", "Department", "Host",
               "Purpose", "Status", "Entry Date", "Entry Time",
               "Exit Time", "Duration"]
    rows = [headers]
    status_map = {
        "Approved": colors.HexColor("#dcfce7"),
        "Pending":  colors.HexColor("#fef9c3"),
        "Rejected": colors.HexColor("#fee2e2"),
        "Checked Out": colors.HexColor("#dbeafe"),
    }
    row_colors = []
    for i, v in enumerate(visitors, 1):
        rows.append([
            str(i),
            v.get("visitor_number", ""),
            v.get("full_name", ""),
            v.get("department", ""),
            v.get("person_to_visit", ""),
            v.get("purpose", "")[:20],
            v.get("status", ""),
            v.get("entry_date", ""),
            v.get("entry_time", ""),
            v.get("exit_time", ""),
            v.get("visit_duration", ""),
        ])
        row_colors.append(status_map.get(v.get("status", ""), colors.white))

    col_widths = [8*mm, 22*mm, 40*mm, 28*mm, 38*mm, 36*mm, 22*mm,
                  22*mm, 18*mm, 18*mm, 16*mm]
    table = Table(rows, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("BACKGROUND",  (0, 0), (-1, 0), navy),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE",    (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0), (-1, -1), 4),
    ]
    # Status row coloring
    for i, rc in enumerate(row_colors, 1):
        style_cmds.append(("BACKGROUND", (6, i), (6, i), rc))

    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    doc.build(story)
    return buf.getvalue()
